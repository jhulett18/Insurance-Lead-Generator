# TODO

from openpyxl import Workbook
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import pandas as pd

import time
import os

# Shorthanding
url = ("Feb.xlsx") 
wb = load_workbook(filename=url)
# Init
ws = wb.active


# ChromeDriver Installation
s = Service(r"C:/Users/JT/Documents/chromedriver.exe")
page = webdriver.Chrome(service=s)


# start
def getInput():
    
    # global variables
    
    global SCISDescList 
    start_list = 0 # Subtract 1 from your starting point to get true starting point in CODE
                        # When copying from tempsave to new_month EXCEL move up 1 square, thats where the excel starts adding desc
    slots = start_list
    end_list = 775 # end of Excel sheet
    
    
    # SCISDescList = [] SCIS INITIAL STATE (USE THIS WHEN YOU WANT TO START FROM 0)
    SCISDescList = [""]  * slots # Has to be in line with start_list to grab descriptions 

    
    print(len(SCISDescList))
    loopCount = start_list # Keep track of Count starting with start_list
   
    # Getting Class Code column
    df_temp = pd.read_excel(url)
    
    classCodeList = (df_temp["Class Code"])
    conversion = classCodeList.tolist()
    
    # Initial Loop
    # Use slicing to specify where to begin the for loop if it crashes
    # Example: for i in conversion[12:]
    
    for i in conversion[start_list:end_list]: 
        # grab site  
        page.get("https://www.insurancexdate.com/class")
        
        # Start
        selectState()
        classCode(i) # This takes in classCode to be evaluated        
        clickButton()
        selectResult()

        # Adds classCode to list TIP: Check getAgentInfo for return value
        SCISDescList.append(getAgentInfo()) # Adds classCode to list        
        

        loopCount = loopCount + 1

        # subtracting to keep index in range (1 - 0 = 0)
        grabDescItem = loopCount - 1

        # Assigning Cell values 
        row = 'K'
        number = str(loopCount)
        cell = row + number
        
        print('\n\n----------------FEEDBACK-----------------------\n') 
        
        print('SCIS #: ' + str(i))
        print('Loop Iteration #: ' + str(loopCount))
        print("Cell #: " + cell)
        print(SCISDescList[grabDescItem])

        # Inserts description into excel cell 
        ws[cell] = SCISDescList[grabDescItem] 

        print('\n-----------------------END FEEDBACK ----------------\n')

        
        # Save after every iteration 
        wb.save("TempSave.xlsx")
        print("Saved Successfully!\n")
        
            
    # Finish 
    print("All processess have been complete: " + loopCount + "entries")
    page.quit()


# State
def selectState():

    try:
        element = WebDriverWait(page, 10).until(
            EC.presence_of_element_located((By.ID, "state"))
        )
    except:
        print("State: Good ✅ ")

    finally:
        # FL value - Florida 
        searchBy = Select(page.find_element_by_id('state'))
        searchBy.select_by_value("FL")
        time.sleep(1)
        
# Inputting classCode into search field
def classCode(i):
    
    xpath = '/html/body/div[1]/div/div[2]/div[1]/form/div/input'
    
    try:
        element = WebDriverWait(page, 10).until(
            EC.presence_of_element_located((By.xpath, xpath))
        )
    except:
        print("Found Text Field")

    finally:
        # FL value - Florida
        # count = cell("How many class codes?: ")
        searchBy = page.find_element_by_xpath(xpath)
        searchBy.send_keys(i)
        time.sleep(1) 
  
# Click Search               
def clickButton():
    xpath = '//button[@type="submit"]'
    
    try:
        element = WebDriverWait(page, 10).until(
            EC.presence_of_element_located((By.xpath, xpath ))
        )
    except:
        print("Found button")

    finally:
        button = page.find_element_by_xpath(xpath)
        button.click()

# Click First Element
def selectResult():
    getOption = "/html/body/div[1]/div/div[2]/div[1]/div/div/div[2]/div/table/tbody/tr[1]/td[2]/span"
        
    try:
        element = WebDriverWait(page, 10).until(
            EC.presence_of_element_located((By.xpath, getOption))
        )
    except:
        print("found detail screen")

    finally:
        selectResult = page.find_element_by_xpath(getOption)
        selectResult.click()
        time.sleep(1)
 
# Grab SCIS Desc 
def getAgentInfo():
    
    classCodeDesc = "/html/body/div[1]/div/div[2]/div[1]/div[2]/div[2]/div[3]"
    employeeClass = "/html/body/div[1]/div/div[2]/div[1]/div[2]/h1"
    # carrierTitle = "/html/body/div[1]/div/div[2]/div[1]/div[2]/div[2]/div[4]/h4"
    # carrierList = "/html/body/div[1]/div/div[2]/div[1]/div[2]/div[2]/div[4]"
    
    try:
        element = WebDriverWait(page, 10).until(
            EC.presence_of_element_located((By.xpath, employeeClass))
        )
    except:
        print("found Agent xpath")

    finally: 
        
        # Setting global variables
        global employeeResult, classCodeResult
        
        # Finding variables

        employeeResult = page.find_element_by_xpath(employeeClass).text
        classCodeResult = page.find_element_by_xpath(classCodeDesc).text
        
        result = ('Employee Section: ' + employeeResult + '\n' + 'Class Code Section: ' + classCodeResult)
        resultFormatted = ('++++++++++++\n\n' + result + "\n")
        
        return resultFormatted 
        
        
# main program
def scisFinder():
    
    getInput()   

scisFinder()